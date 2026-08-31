from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

def export_tender_to_excel(tender_data: dict, items: list[dict]) -> BytesIO:
    """Экспортирует тендер в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Тендер"
    
    # Заголовок
    ws['A1'] = tender_data.get('name', 'Тендер')
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')
    
    ws['A2'] = f"Номер: {tender_data.get('number', 'Не указан')}"
    ws['A3'] = f"Регион: {tender_data.get('region', 'Не указан')}"
    ws['A4'] = f"Дата создания: {tender_data.get('created_at', '')}"
    
    # Шапка таблицы
    headers = ['№', 'Наименование', 'Количество', 'Ед.изм.', 'Производитель', 'Модель', 'Цена']
    ws.append([])
    ws.append(headers)
    
    header_row = ws[6]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for item in items:
        ws.append([
            item.get('position_number', ''),
            item.get('name', ''),
            item.get('quantity', 1),
            item.get('unit', 'шт'),
            item.get('manufacturer', ''),
            item.get('model', ''),
            item.get('price', '')
        ])
    
    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_cp_to_excel(cp_data: dict) -> BytesIO:
    """Экспортирует КП в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "КП"
    
    # Заголовок
    ws['A1'] = "КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    row = 3
    ws[f'A{row}'] = f"Поставщик: {cp_data.get('supplier_name', '')}"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = f"ИНН: {cp_data.get('supplier_inn', '')}"
    row += 1
    ws[f'A{row}'] = f"Контакт: {cp_data.get('contact', '')}"
    row += 2
    
    ws[f'A{row}'] = f"Тендер: {cp_data.get('tender_name', '')}"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = f"Номер: {cp_data.get('tender_number', 'Не указан')}"
    row += 2
    
    # Таблица
    headers = ['№', 'Наименование', 'Производитель', 'Модель', 'Кол-во', 'Ед.', 'Цена', 'Сумма']
    ws.append(headers)
    header_row = ws[row]
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    for item in cp_data.get('items', []):
        ws.append([
            item.get('position', ''),
            item.get('name', ''),
            item.get('manufacturer', ''),
            item.get('model', ''),
            item.get('quantity', 1),
            item.get('unit', 'шт'),
            item.get('price', 0),
            item.get('sum', 0)
        ])
        row += 1
    
    # Итого
    row += 1
    ws[f'G{row}'] = "Итого:"
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = cp_data.get('total', 0)
    ws[f'H{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'G{row}'] = "НДС 20%:"
    ws[f'H{row}'] = cp_data.get('vat', 0)
    
    row += 1
    ws[f'G{row}'] = "Всего с НДС:"
    ws[f'G{row}'].font = Font(bold=True, size=12)
    ws[f'H{row}'] = cp_data.get('total_with_vat', 0)
    ws[f'H{row}'].font = Font(bold=True, size=12)
    
    row += 2
    ws[f'A{row}'] = f"Срок поставки: {cp_data.get('delivery_days', '')} дней (до {cp_data.get('delivery_date', '')})"
    row += 1
    ws[f'A{row}'] = f"Условия оплаты: {cp_data.get('payment_terms', '')}"
    row += 1
    ws[f'A{row}'] = f"Гарантия: {cp_data.get('warranty', '')}"
    
    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
